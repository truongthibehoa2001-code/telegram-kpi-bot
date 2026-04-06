from telegram.ext import Application, MessageHandler, filters
from openpyxl import Workbook, load_workbook
import os

TOKEN = "8587229409:AAHtZp84s_4MwDG4WP1d7l5poD0s02wOBw8"
FILE_NAME = "nhanvien_note.xlsx"


def init_excel():
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.title = "DATA"
        ws.append([
            "Tên Nhân Viên",
            "Ngày",
            "Loại Tài Khoản",
            "Số Lượng",
            "Danh Sách ID"
        ])
        wb.save(FILE_NAME)


async def note_handler(update, context):
    text = update.message.text
    if not text.startswith("/note"):
        return

    lines = text.split("\n")

    try:
        name = lines[1].split(":")[1].strip()
        date = lines[2].split(":")[1].strip()
        acc_type = lines[3].split(":")[1].strip()
        quantity = lines[4].split(":")[1].strip()
        ids = lines[6:]

        wb = load_workbook(FILE_NAME)
        ws = wb["DATA"]

        ws.append([
            name,
            date,
            acc_type,
            quantity,
            "\n".join(ids)
        ])

        wb.save(FILE_NAME)

        await update.message.reply_text(
            f"✅ {name} đã note thành công\n📦 +{quantity} {acc_type}"
        )

    except:
        await update.message.reply_text("❌ Sai format /note")


def main():
    init_excel()

    app = Application.builder().token(TOKEN).build()
    app.add_handler(MessageHandler(filters.TEXT, note_handler))

    print("🤖 BOT đang chạy...")
    app.run_polling()


if __name__ == "__main__":
    main()